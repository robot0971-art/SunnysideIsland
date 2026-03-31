import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

output_dir = "GameData_Excel"
os.makedirs(output_dir, exist_ok=True)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_sheet(ws, headers, data):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

# 9. Quests.xlsx
wb = openpyxl.Workbook()

# Main Quests sheet
ws = wb.active
ws.title = "MainQuests"
quest_headers = ["QuestID", "QuestName", "QuestType", "Chapter", "Description", "Objectives", "Rewards", "Prerequisites"]
quest_data = [
    ["quest_crash", "불시착", "Main", 1, "텐트를 건설하세요", "build:tent", "basic_tool_set", ""],
    ["quest_first_meal", "첫 식사", "Main", 1, "낚시로 생선 3마리를 잡으세요", "fish:3", "fishing_rod_upgrade", "quest_crash"],
    ["quest_safe_place", "안전한 곳", "Main", 1, "오두막을 완성하세요", "build:hut", "bed_recipe", "quest_first_meal"],
    ["quest_farmer_start", "농부의 시작", "Main", 2, "첫 작물을 수확하세요", "harvest:1", "seeds:20", "quest_safe_place"],
    ["quest_new_neighbor", "새로운 이웃", "Main", 2, "첫 주민을 유치하세요", "recruit:1", "resident_furniture_set", "quest_farmer_start"],
    ["quest_goblin_threat", "고블린의 위협", "Main", 2, "고블린 5마리를 처치하세요", "defeat:goblin:5", "combat_tool_recipe", "quest_new_neighbor"],
    ["quest_discover_village", "고블린 마을 발견", "Main", 3, "고블린 마을을 탐색하세요", "explore:goblin_village", "goblin_translator", "quest_goblin_threat"],
    ["quest_peace", "평화의 조건", "Main", 3, "고블린 족장을 처치하거나 교역 협정을 체결하세요", "defeat:goblin_chief OR trade:goblin", "goblin_alliance OR treasure", "quest_discover_village"],
    ["quest_first_tourist", "첫 관광객", "Main", 3, "관광객 10명을 유치하세요", "attract_tourist:10", "pier_upgrade", "quest_peace"],
    ["quest_merchant_path", "상인의 길", "Main", 4, "상점 5개를 건설하세요", "build:shop:5", "market_unlock", "quest_first_tourist"],
    ["quest_festival", "축제의 날", "Main", 4, "첫 축제를 개최하세요", "host:festival", "reputation_boost", "quest_merchant_path"],
    ["quest_resort_complete", "리조트 완성", "Main", 4, "리조트 호텔을 건설하세요", "build:resort_hotel", "game_clear,free_mode", "quest_festival"],
]
create_sheet(ws, quest_headers, quest_data)

# Sub Quests sheet
ws2 = wb.create_sheet("SubQuests")
sub_quest_headers = ["QuestID", "QuestName", "QuestType", "Chapter", "Description", "Objectives", "Rewards", "Prerequisites"]
sub_quest_data = [
    ["quest_fisherman_john", "전설의 물고기", "Sub", 0, "낚시꾼 존의 요청", "catch:golden_fish", "legendary_fishing_rod", ""],
    ["quest_blacksmith", "희귀 광석", "Sub", 0, "대장장이의 요청", "mine:mithril:10", "steel_armor", ""],
    ["quest_chef", "미식 대회", "Sub", 0, "요리사의 요청", "cook:special:3", "chef_resident", ""],
    ["quest_goblin_merchant", "물물교환", "Sub", 0, "고블린 상인의 요청", "trade:human_items:10", "goblin_special_goods", ""],
    ["quest_explorer", "잃어버린 유물", "Sub", 0, "탐험가의 요청", "explore:ancient_ruins", "map,treasure", ""],
]
create_sheet(ws2, sub_quest_headers, sub_quest_data)

wb.save(os.path.join(output_dir, "09_Quests.xlsx"))
print(f"Created: {os.path.join(output_dir, '09_Quests.xlsx')}")

# 10. Skills.xlsx
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Skills"
skill_headers = ["SkillID", "SkillTree", "Level", "SkillName", "Effect", "Value"]
skill_data = [
    # Gathering
    ["skill_gather_1", "Gathering", 1, "벌목 속도 증가", "벌목 속도 +10%", 10],
    ["skill_gather_2", "Gathering", 2, "채광 속도 증가", "채광 속도 +10%", 10],
    ["skill_gather_3", "Gathering", 3, "희귀 자원 발견", "희귀 자원 발견 +5%", 5],
    ["skill_gather_4", "Gathering", 4, "내구도 절약", "내구도 소모 -20%", 20],
    ["skill_gather_5", "Gathering", 5, "더블 채집", "한 번에 2개 채집 가능", 2],
    # Farming
    ["skill_farm_1", "Farming", 1, "성장 가속", "성장 속도 +10%", 10],
    ["skill_farm_2", "Farming", 2, "수확량 증가", "수확량 +1", 1],
    ["skill_farm_3", "Farming", 3, "물 주기 간격 감소", "물 주기 필요 주기 -1일", 1],
    ["skill_farm_4", "Farming", 4, "품질 상승", "품질 상승 확률 +20%", 20],
    ["skill_farm_5", "Farming", 5, "계절 무관", "계절 무관 작물 재배", 1],
    # Fishing
    ["skill_fish_1", "Fishing", 1, "빠른 미끼", "미끼 대기 시간 -20%", 20],
    ["skill_fish_2", "Fishing", 2, "쉬운 낚시", "미니게임 난이도 -10%", 10],
    ["skill_fish_3", "Fishing", 3, "희귀 물고기", "희귀 물고기 확률 +10%", 10],
    ["skill_fish_4", "Fishing", 4, "낚싯대 수명", "낚싯대 내구도 소모 -30%", 30],
    ["skill_fish_5", "Fishing", 5, "대물 낚시", "물고기 크기/품질 상승", 1],
    # Combat
    ["skill_combat_1", "Combat", 1, "공격력 증가", "공격력 +10%", 10],
    ["skill_combat_2", "Combat", 2, "방어력 증가", "방어력 +10%", 10],
    ["skill_combat_3", "Combat", 3, "치명타", "치명타 확률 +5%", 5],
    ["skill_combat_4", "Combat", 4, "스태미나 절약", "스태미나 소모 -20%", 20],
    ["skill_combat_5", "Combat", 5, "회전베기", "특수 공격 해금 (회전베기)", 1],
    # ConstructionEconomy
    ["skill_build_1", "ConstructionEconomy", 1, "빠른 건설", "건설 속도 +20%", 20],
    ["skill_build_2", "ConstructionEconomy", 2, "건설 비용 절감", "건설 비용 -10%", 10],
    ["skill_build_3", "ConstructionEconomy", 3, "판매 가격 증가", "판매 가격 +10%", 10],
    ["skill_build_4", "ConstructionEconomy", 4, "주민 생산성", "주민 생산성 +15%", 15],
    ["skill_build_5", "ConstructionEconomy", 5, "관광객 유치", "관광객 유치 +25%", 25],
]
create_sheet(ws, skill_headers, skill_data)
wb.save(os.path.join(output_dir, "10_Skills.xlsx"))
print(f"Created: {os.path.join(output_dir, '10_Skills.xlsx')}")

# 11. Events.xlsx
wb = openpyxl.Workbook()

# Weekly Events sheet
ws = wb.active
ws.title = "WeeklyEvents"
event_headers = ["EventID", "EventName", "EventType", "TriggerCondition", "Probability", "Effect", "Duration"]
event_data = [
    ["event_monday", "채집의 날", "Weekly", "DayOfWeek=Monday", 100, "채집 XP 2배", 1],
    ["event_wednesday", "낚시 대회", "Weekly", "DayOfWeek=Wednesday", 100, "특별 물고기 출현", 1],
    ["event_friday", "시장의 날", "Weekly", "DayOfWeek=Friday", 100, "모든 가격 할인 20%", 1],
    ["event_sunday", "쉬는 날", "Weekly", "DayOfWeek=Sunday", 100, "모든 NPC 휴식", 1],
]
create_sheet(ws, event_headers, event_data)

# Seasonal Events sheet
ws2 = wb.create_sheet("SeasonalEvents")
seasonal_data = [
    ["event_spring", "식목일", "Seasonal", "Season=Spring", 100, "나무 무료 제공", 1],
    ["event_summer", "해변 축제", "Seasonal", "Season=Summer", 100, "관광객 2배", 7],
    ["event_autumn", "추수제", "Seasonal", "Season=Autumn", 100, "농작물 수확량 2배", 7],
    ["event_winter", "눈축제", "Seasonal", "Season=Winter", 100, "겨울 한정 아이템", 7],
]
create_sheet(ws2, event_headers, seasonal_data)

# Random Events sheet
ws3 = wb.create_sheet("RandomEvents")
random_data = [
    ["event_merchant_visit", "상인 방문", "Random", "DailyCheck", 10, "희귀 아이템 판매", 1],
    ["event_goblin_attack", "고블린 습격", "Random", "DailyCheck", 5, "마을 방어전", 0],
    ["event_treasure", "보물 발견", "Random", "DailyCheck", 2, "랜덤 보물 상자", 0],
    ["event_rainbow", "무지개", "Random", "DailyCheck", 1, "행운 최대", 1],
    ["event_ghost", "유령 출몰", "Random", "NightCheck", 3, "특별 몬스터 출현", 0],
    ["event_bumper_crop", "풍작", "Random", "DailyCheck", 5, "작물 성장 속도 2배", 3],
    ["event_drought", "가뭄", "Random", "DailyCheck", 3, "물 주기 필요 증가", 3],
]
create_sheet(ws3, event_headers, random_data)

wb.save(os.path.join(output_dir, "11_Events.xlsx"))
print(f"Created: {os.path.join(output_dir, '11_Events.xlsx')}")

print("\n✅ Excel 파일 생성 완료 (Part 3)")
