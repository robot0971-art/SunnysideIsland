import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

output_dir = "GameData_Excel"
os.makedirs(output_dir, exist_ok=True)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_sheet(ws, headers, data):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

# 6. Buildings.xlsx
wb = openpyxl.Workbook()

# Buildings sheet
ws = wb.active
ws.title = "Buildings"
building_headers = ["BuildingID", "BuildingName", "BuildingType", "SizeX", "SizeY", "BuildTime", "RequiredResources", "EffectDescription"]
building_data = [
    ["tent", "텐트", "Residential", 2, 2, 0, "cloth:3,wood:5", "임시 거처"],
    ["hut", "오두막", "Residential", 3, 3, 60, "wood:30,stone:10", "기본 주거"],
    ["house", "집", "Residential", 4, 4, 120, "wood:50,stone:30,iron:10", "주민 수용 가능"],
    ["large_house", "큰 집", "Residential", 5, 5, 240, "wood:100,stone:60,iron:30", "2인 주거"],
    ["mansion", "저택", "Residential", 6, 6, 480, "stone:200,iron:50,glass:20", "고급 거처"],
    ["farm_plot", "밭", "Agriculture", 2, 2, 0, "wood:10", "4칸 작물 재배"],
    ["large_farm", "큰 밭", "Agriculture", 4, 4, 0, "wood:30,stone:10", "16칸 작물 재배"],
    ["greenhouse", "비닐하우스", "Agriculture", 6, 4, 0, "glass:20,iron:15", "계절 무관, 성장 +20%"],
    ["barn", "축사", "Agriculture", 4, 4, 0, "wood:40,stone:20", "동물 사육"],
    ["storage", "저장고", "Agriculture", 3, 3, 0, "wood:50,iron:10", "식량 보관"],
    ["windmill", "풍차", "Agriculture", 2, 2, 0, "wood:30,iron:20", "자동 물 공급"],
    ["workbench", "작업대", "Production", 2, 2, 0, "wood:20", "기본 제작"],
    ["furnace", "화덕", "Production", 2, 2, 0, "stone:30", "요리, 광석 용해"],
    ["blast_furnace", "용광로", "Production", 3, 3, 0, "stone:50,iron:10", "고급 제련"],
]
create_sheet(ws, building_headers, building_data)

# Commercial Buildings sheet
ws2 = wb.create_sheet("Commercial")
commercial_headers = ["BuildingID", "DailyIncomeMin", "DailyIncomeMax", "SpecialEffect"]
commercial_data = [
    ["market_stall", 50, 100, "기본 거래"],
    ["grocery", 150, 300, "농작물 판매 +20%"],
    ["blacksmith", 200, 400, "장비 수리/강화"],
    ["fishing_shop", 100, 250, "낚시 도구 판매"],
    ["gift_shop", 250, 500, "관광객 전용"],
    ["restaurant", 400, 800, "요리 판매"],
    ["inn", 500, 1000, "관광객 숙박"],
    ["market", 1000, 2000, "모든 NPC 거래"],
]
create_sheet(ws2, commercial_headers, commercial_data)

# Tourist Buildings sheet
ws3 = wb.create_sheet("Tourist")
tourist_headers = ["BuildingID", "TouristIncrease"]
tourist_building_data = [
    ["pier", 0],
    ["lighthouse", 10],
    ["plaza", 15],
    ["park", 20],
    ["museum", 25],
    ["festival_ground", 30],
    ["hot_spring", 35],
    ["resort_hotel", 50],
]
create_sheet(ws3, tourist_headers, tourist_building_data)

wb.save(os.path.join(output_dir, "06_Buildings.xlsx"))
print(f"Created: {os.path.join(output_dir, '06_Buildings.xlsx')}")

# 7. NPCs.xlsx
wb = openpyxl.Workbook()

# Residents sheet
ws = wb.active
ws.title = "Residents"
resident_headers = ["ResidentID", "ResidentType", "Function", "DailyWage", "SpecialAbility", "RequiredCondition"]
resident_data = [
    ["merchant", "Merchant", "자동 판매, 재고 관리", 50, "가격 협상 +10%", "상점 1개"],
    ["blacksmith_npc", "Blacksmith", "장비 수리, 강화", 80, "수리비 -20%", "대장간"],
    ["chef", "Chef", "음식 대량 제작", 60, "요리 품질 상승", "식당"],
    ["farmer", "Farmer", "농사 자동화", 40, "작물 성장 +10%", "밭 8칸 이상"],
    ["carpenter", "Carpenter", "가구 제작", 50, "건설 속도 +20%", "공방"],
    ["guard", "Guard", "마을 방어", 70, "치안 레벨 상승", "고블린 위협 있음"],
    ["researcher", "Researcher", "마법 아이템 개발", 100, "새로운 레시피", "마법 연구소"],
    ["guide", "Guide", "관광객 안내", 60, "관광객 만족도 +20%", "관광객 10명 이상"],
]
create_sheet(ws, resident_headers, resident_data)

# Tourist Types sheet
ws2 = wb.create_sheet("TouristTypes")
tourist_type_headers = ["TypeID", "TypeName", "Ratio", "PreferredFacilities", "SpendingMin", "SpendingMax"]
tourist_type_data = [
    ["normal", "일반", 50, "식당,기념품", 200, 300],
    ["premium", "고급", 20, "온천,고급식당", 500, 1000],
    ["explorer", "탐험가", 20, "동굴,등산", 300, 500],
    ["group", "단체", 10, "축제장,광장", 1500, 3000],
]
create_sheet(ws2, tourist_type_headers, tourist_type_data)

wb.save(os.path.join(output_dir, "07_NPCs.xlsx"))
print(f"Created: {os.path.join(output_dir, '07_NPCs.xlsx')}")

# 8. Recipes.xlsx
wb = openpyxl.Workbook()

# Recipes sheet
ws = wb.active
ws.title = "Recipes"
recipe_headers = ["RecipeID", "RecipeName", "ResultItemID", "Ingredients", "HungerRestore", "AdditionalEffect", "CookTime"]
recipe_data = [
    ["grilled_fish", "구운 생선", "grilled_fish_item", "fish:1", 20, "", 30],
    ["fish_roast", "생선구이", "fish_roast_item", "fish:2,salt:1", 45, "체력 +5", 60],
    ["stew", "스튜", "stew_item", "meat:2,potato:2,carrot:1", 60, "체력 +10", 120],
    ["salad", "샐러드", "salad_item", "vegetable:3", 30, "스태미나 +20", 30],
    ["bread", "빵", "bread_item", "flour:2", 40, "", 180],
    ["sandwich", "샌드위치", "sandwich_item", "bread:2,meat:1,vegetable:1", 50, "", 60],
    ["pie", "파이", "pie_item", "flour:2,fruit:3", 55, "기분 상승", 240],
    ["steak", "스테이크", "steak_item", "premium_meat:2", 80, "공격력 +10% (10분)", 120],
    ["seafood_pasta", "해산물 파스타", "seafood_pasta_item", "fish:2,flour:2", 70, "모든 능력치 +5", 180],
    ["royal_platter", "로얄 플래터", "royal_platter_item", "fish:3,meat:3,vegetable:3", 100, "모든 능력치 +15 (30분)", 300],
]
create_sheet(ws, recipe_headers, recipe_data)

# Crafting Recipes sheet
ws2 = wb.create_sheet("Crafting")
crafting_headers = ["RecipeID", "Category", "ResultItemID", "ResultAmount", "Ingredients", "RequiredTool"]
crafting_data = [
    ["recipe_wood_axe", "Tool", "wood_axe", 1, "wood:10", ""],
    ["recipe_stone_pickaxe", "Tool", "stone_pickaxe", 1, "wood:5,stone:10", ""],
    ["recipe_iron_fishing_rod", "Tool", "iron_fishing_rod", 1, "wood:10,iron:5", ""],
    ["recipe_iron_sword", "Equipment", "iron_sword", 1, "iron:10,wood:5", ""],
    ["recipe_wood_bow", "Equipment", "wood_bow", 1, "wood:15,rope:5", ""],
    ["recipe_wood_plank", "Construction", "wood_plank", 2, "wood:2", ""],
    ["recipe_stone_brick", "Construction", "stone_brick", 2, "stone:2", ""],
    ["recipe_iron_ingot", "Processing", "iron_ingot", 1, "iron_ore:3", "furnace"],
    ["recipe_steel", "Processing", "steel", 1, "iron_ingot:2,coal:1", "blast_furnace"],
    ["recipe_bed", "Furniture", "bed", 1, "wood:20,cloth:10", "workbench"],
    ["recipe_desk", "Furniture", "desk", 1, "wood:15", "workbench"],
]
create_sheet(ws2, crafting_headers, crafting_data)

wb.save(os.path.join(output_dir, "08_Recipes.xlsx"))
print(f"Created: {os.path.join(output_dir, '08_Recipes.xlsx')}")

print("\n✅ Excel 파일 생성 완료 (Part 2)")
