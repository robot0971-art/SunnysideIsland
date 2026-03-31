import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

output_dir = "GameData_Excel"
os.makedirs(output_dir, exist_ok=True)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_sheet(ws, headers, data):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

# 12. Achievements.xlsx
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Achievements"
achievement_headers = ["AchievementID", "AchievementName", "Category", "RequirementType", "TargetID", "TargetAmount", "Reward"]
achievement_data = [
    # Gathering
    ["ach_woodcutter", "나무꾼", "Gathering", "Collect", "wood", 1000, "title:나무꾼"],
    ["ach_miner", "광부", "Gathering", "Collect", "iron_ore", 500, "title:광부"],
    ["ach_collector", "수집가", "Gathering", "Own", "all_resources", 999, "title:수집가"],
    # Farming
    ["ach_farmer", "농부", "Farming", "Collect", "crop", 1, ""],
    ["ach_big_farmer", "대농장주", "Farming", "Collect", "crop", 1000, "title:대농장주"],
    ["ach_gardener", "원예가", "Farming", "Collect", "all_crops", 1, "title:원예가"],
    # Fishing
    ["ach_fisherman", "낚시꾼", "Fishing", "Collect", "fish", 1, ""],
    ["ach_fishing_king", "낚시왕", "Fishing", "Collect", "fish", 1000, "title:낚시왕"],
    ["ach_legendary_fisherman", "전설의 낚시꾼", "Fishing", "Collect", "golden_fish", 1, "title:전설의 낚시꾼"],
    # Combat
    ["ach_warrior", "전사", "Combat", "Defeat", "monster", 1, ""],
    ["ach_goblin_killer", "고블린 킬러", "Combat", "Defeat", "goblin", 100, "title:고블린 킬러"],
    ["ach_boss_slayer", "보스 슬레이어", "Combat", "Defeat", "all_bosses", 1, "title:보스 슬레이어"],
    # Building
    ["ach_builder", "건축가", "Building", "Build", "building", 1, ""],
    ["ach_developer", "도시 개발자", "Building", "Build", "building", 50, "title:도시 개발자"],
    ["ach_resort_king", "리조트 왕", "Building", "Build", "all_buildings", 1, "title:리조트 왕"],
    # Economy
    ["ach_rich", "부자", "Economy", "Own", "coin", 10000, "title:부자"],
    ["ach_mogul", "거물", "Economy", "Own", "coin", 100000, "title:거물"],
    ["ach_president", "경제 대통령", "Economy", "Own", "coin", 1000000, "title:경제 대통령"],
    # Special
    ["ach_perfectionist", "완벽주의자", "Special", "Achieve", "all_achievements", 1, "title:완벽주의자"],
    ["ach_speedrunner", "스피드러너", "Special", "Clear", "game", 120, "title:스피드러너"],
    ["ach_irman", "철인", "Special", "Clear", "game_no_combat", 1, "title:철인"],
]
create_sheet(ws, achievement_headers, achievement_data)
wb.save(os.path.join(output_dir, "12_Achievements.xlsx"))
print(f"Created: {os.path.join(output_dir, '12_Achievements.xlsx')}")

# 13. World.xlsx
wb = openpyxl.Workbook()

# TimeOfDay sheet
ws = wb.active
ws.title = "TimeOfDay"
time_headers = ["TimeSlot", "StartHour", "EndHour", "Feature", "RecommendedActivity"]
time_data = [
    ["Dawn", 4, 6, "어두움, 몬스터 활동 증가", "-"],
    ["Morning", 6, 9, "밝아짐, NPC 기상", "농사, 채집"],
    ["Noon", 9, 12, "활동 최적기", "건설, 탐험"],
    ["Afternoon", 12, 14, "NPC 휴식 시간", "낚시, 사냥"],
    ["Evening", 14, 18, "황금 시간", "전투, 채집"],
    ["Night", 18, 21, "어두워짐", "상점 운영, 요리"],
    ["LateNight", 21, 4, "어두움, 몬스터 활동 최대", "휴식, 동굴 탐험"],
]
create_sheet(ws, time_headers, time_data)

# Weather sheet
ws2 = wb.create_sheet("Weather")
weather_headers = ["WeatherType", "Probability", "Effect"]
weather_data = [
    ["Sunny", 50, "기본 상태, 모든 활동 가능"],
    ["Cloudy", 20, "낚시 성공률 +10%"],
    ["Rainy", 15, "농사 자동 물주기, 채집 불편"],
    ["Stormy", 10, "외부 활동 위험, 실내 권장"],
    ["Rainbow", 5, "희귀, 행운 +20, 관광객 증가"],
]
create_sheet(ws2, weather_headers, weather_data)

# ResourceSpawn sheet
ws3 = wb.create_sheet("ResourceSpawn")
resource_headers = ["ResourceID", "ResourceName", "RegenDays", "MaxQuantity", "SpawnAreas"]
resource_data = [
    ["tree", "나무", 3, 100, "Forest,Mountain,Beach"],
    ["stone_ore", "돌", 5, 50, "Mountain,Cave"],
    ["iron_ore", "철광석", 7, 30, "Mountain,Cave"],
    ["mithril_ore", "미스릴", 14, 10, "Cave"],
    ["herb", "약초", 2, 20, "Forest"],
    ["mushroom", "버섯", 1, 30, "Forest,Cave"],
]
create_sheet(ws3, resource_headers, resource_data)

# Areas sheet
ws4 = wb.create_sheet("Areas")
area_headers = ["AreaID", "AreaName", "AreaType", "Difficulty", "AvailableResources", "AvailableAnimals", "AvailableMonsters"]
area_data = [
    ["beach", "해변", "Beach", 1, "tree", "", ""],
    ["forest", "숲", "Forest", 2, "tree,herb,mushroom", "rabbit,deer,fox", "goblin_farmer"],
    ["mountain", "산", "Mountain", 3, "tree,stone_ore,iron_ore", "boar,wolf", "goblin_soldier,goblin_archer"],
    ["ocean", "바다", "Ocean", 1, "", "", ""],
    ["cave", "동굴", "Cave", 4, "stone_ore,iron_ore,mithril_ore,mushroom", "", "goblin_warrior,skeleton"],
    ["goblin_village", "고블린 마을", "GoblinVillage", 5, "", "", "goblin_chief"],
]
create_sheet(ws4, area_headers, area_data)

wb.save(os.path.join(output_dir, "13_World.xlsx"))
print(f"Created: {os.path.join(output_dir, '13_World.xlsx')}")

# 14. Shops.xlsx
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ShopItems"
shop_headers = ["ShopItemID", "ItemID", "BuyPrice", "SellPrice", "StockAmount", "RestockInterval"]
shop_data = [
    ["shop_wood", "wood", 5, 2, -1, 1],
    ["shop_stone", "stone", 8, 3, -1, 1],
    ["shop_iron_ore", "iron_ore", 30, 15, 50, 3],
    ["shop_carrot_seed", "carrot_seed", 30, 0, -1, 1],
    ["shop_tomato_seed", "tomato_seed", 45, 0, -1, 1],
    ["shop_wood_axe", "wood_axe", 150, 50, 5, 7],
    ["shop_iron_sword", "iron_sword", 500, 150, 3, 7],
    ["shop_healing_potion", "healing_potion", 100, 20, 10, 3],
]
create_sheet(ws, shop_headers, shop_data)
wb.save(os.path.join(output_dir, "14_Shops.xlsx"))
print(f"Created: {os.path.join(output_dir, '14_Shops.xlsx')}")

print("\n✅ 모든 Excel 파일 생성 완료!")
print(f"📁 저장 위치: {os.path.abspath(output_dir)}")
print("\n생성된 파일 목록:")
for f in sorted(os.listdir(output_dir)):
    if f.endswith('.xlsx'):
        print(f"  - {f}")
