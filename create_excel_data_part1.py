import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Create output directory
output_dir = "GameData_Excel"
os.makedirs(output_dir, exist_ok=True)

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_workbook(filename, sheet_name, headers, sample_data=None):
    """Create an Excel workbook with headers and optional sample data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add sample data if provided
    if sample_data:
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"Created: {filepath}")
    return filepath

# 1. Items.xlsx
items_headers = ["ItemID", "ItemName", "ItemType", "MaxStack", "BaseValue", "CanSell", "Description", "IconPath"]
items_data = [
    ["wood", "나무", "Material", 999, 2, True, "기본적인 목재입니다.", "Items/wood"],
    ["stone", "돌", "Material", 999, 3, True, "기본적인 돌입니다.", "Items/stone"],
    ["iron_ore", "철광석", "Material", 999, 15, True, "철을 추출할 수 있는 광석입니다.", "Items/iron_ore"],
    ["carrot", "당근", "Consumable", 99, 45, True, "신선한 당근입니다.", "Items/carrot"],
    ["wood_axe", "나무 도끼", "Tool", 1, 100, True, "나무를 벨 수 있는 기본 도구입니다.", "Items/wood_axe"],
]
create_workbook("01_Items.xlsx", "Items", items_headers, items_data)

# 2. Crops.xlsx
crops_headers = ["CropID", "CropName", "GrowthDays", "YieldAmount", "Seasons", "BuyPrice", "SellPrice", "SpecialEffect", "SeedItemID", "CropItemID"]
crops_data = [
    ["carrot", "당근", 3, 3, "봄,가을", 20, 45, "시력 상승", "carrot_seed", "carrot"],
    ["tomato", "토마토", 4, 4, "여름", 30, 60, "반복 수확", "tomato_seed", "tomato"],
    ["potato", "감자", 4, 5, "봄,가을", 25, 55, "배고픔 +30", "potato_seed", "potato"],
    ["corn", "옥수수", 5, 3, "여름", 40, 80, "동물 사료", "corn_seed", "corn"],
    ["watermelon", "수박", 6, 1, "여름", 60, 150, "허기 +50", "watermelon_seed", "watermelon"],
    ["chili", "고추", 4, 6, "여름,가을", 35, 70, "스태미나 +20", "chili_seed", "chili"],
    ["pumpkin", "호박", 5, 2, "가을", 50, 120, "할로윈 이벤트", "pumpkin_seed", "pumpkin"],
    ["wheat", "밀", 4, 4, "봄,가을", 20, 40, "빵 제작", "wheat_seed", "wheat"],
    ["sunflower", "해바라기", 5, 1, "여름", 80, 200, "기분 상승", "sunflower_seed", "sunflower"],
    ["eggplant", "가지", 4, 3, "여름", 45, 90, "고급 요리", "eggplant_seed", "eggplant"],
    ["onion", "양파", 3, 4, "봄", 25, 50, "요리 재료", "onion_seed", "onion"],
]
create_workbook("02_Crops.xlsx", "Crops", crops_headers, crops_data)

# 3. Fishing.xlsx - Fish
fish_headers = ["FishID", "FishName", "Grade", "Location", "TimeCondition", "Difficulty", "SellPrice", "HungerRestore", "ItemID"]
fish_data = [
    ["bass", "농어", 1, "바다", "전시간", "Easy", 30, 15, "fish_bass"],
    ["carp", "잉어", 1, "호수", "전시간", "Easy", 25, 12, "fish_carp"],
    ["catfish", "메기", 1, "강", "밤", "Easy", 35, 18, "fish_catfish"],
    ["tuna", "참치", 2, "바다 깊은 곳", "낮", "Normal", 80, 30, "fish_tuna"],
    ["salmon", "연어", 2, "폭포", "아침", "Normal", 100, 35, "fish_salmon"],
    ["swordfish", "황새치", 3, "바다", "낮", "Hard", 250, 50, "fish_swordfish"],
    ["shark", "상어", 3, "바다 깊은 곳", "밤", "VeryHard", 500, 80, "fish_shark"],
    ["golden_fish", "황금 물고기", 4, "특별 장소", "무지개 날", "Extreme", 2000, 100, "fish_golden"],
]
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Fish"
for col, header in enumerate(fish_headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
for row_idx, row_data in enumerate(fish_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border

# Fishing Rods sheet
ws2 = wb.create_sheet("FishingRods")
fishing_rod_headers = ["RodID", "RodName", "SuccessRate", "Durability", "RareFishBonus", "ItemID"]
fishing_rod_data = [
    ["wooden_rod", "나무 낚싯대", 0, 50, 0, "rod_wooden"],
    ["iron_rod", "철제 낚싯대", 20, 100, 10, "rod_iron"],
    ["gold_rod", "금 낚싯대", 35, 150, 20, "rod_gold"],
    ["legendary_rod", "전설의 낚싯대", 50, 200, 30, "rod_legendary"],
]
for col, header in enumerate(fishing_rod_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
for row_idx, row_data in enumerate(fishing_rod_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border

wb.save(os.path.join(output_dir, "03_Fishing.xlsx"))
print(f"Created: {os.path.join(output_dir, '03_Fishing.xlsx')}")

# 4. Animals.xlsx
animals_headers = ["AnimalID", "AnimalName", "HP", "AttackPower", "Speed", "AIType", "DropItems"]
animals_data = [
    ["rabbit", "토끼", 20, 0, "Fast", "Flee", "meat:1,leather:1"],
    ["deer", "사슴", 40, 0, "VeryFast", "Flee", "meat:2,leather:2"],
    ["fox", "여우", 30, 5, "Fast", "Evasive", "meat:1,leather:1,tail:1"],
    ["boar", "멧돼지", 80, 15, "Normal", "Hostile", "meat:3,leather:2,tusk:2"],
    ["wolf", "늑대", 60, 20, "Fast", "Hostile", "meat:2,leather:2,fang:2"],
    ["bear", "곰", 150, 30, "Slow", "Territorial", "meat:5,leather:4,paw:1"],
]
create_workbook("04_Animals.xlsx", "Animals", animals_headers, animals_data)

# 5. Combat.xlsx - Weapons
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Weapons"
weapon_headers = ["WeaponID", "WeaponName", "AttackPower", "AttackSpeed", "RangeType", "SpecialEffect", "RecipeID", "ItemID"]
weapon_data = [
    ["wood_sword", "나무 검", 10, "Normal", "Melee", "", "recipe_wood_sword", "weapon_wood_sword"],
    ["stone_axe", "돌 도끼", 15, "Slow", "Melee", "벌목 보너스", "recipe_stone_axe", "weapon_stone_axe"],
    ["iron_sword", "철 검", 25, "Fast", "Melee", "", "recipe_iron_sword", "weapon_iron_sword"],
    ["steel_sword", "강철 검", 40, "Fast", "Melee", "", "recipe_steel_sword", "weapon_steel_sword"],
    ["magic_sword", "환검", 60, "VeryFast", "Melee", "마법 데미지", "recipe_magic_sword", "weapon_magic_sword"],
    ["wood_bow", "나무 활", 12, "Normal", "LongRange", "화살 필요", "recipe_wood_bow", "weapon_wood_bow"],
    ["iron_bow", "철 활", 28, "Fast", "LongRange", "관통 효과", "recipe_iron_bow", "weapon_iron_bow"],
    ["magic_staff", "마법 지팡이", 20, "Normal", "MidRange", "마나 소모", "recipe_magic_staff", "weapon_magic_staff"],
]
for col, header in enumerate(weapon_headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
for row_idx, row_data in enumerate(weapon_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border

# Monsters sheet
ws2 = wb.create_sheet("Monsters")
monster_headers = ["MonsterID", "MonsterName", "HP", "AttackPower", "Defense", "Speed", "DropItems", "ExpReward"]
monster_data = [
    ["goblin_farmer", "고블린 농부", 30, 8, 2, "Slow", "meat,goblin_leather", 10],
    ["goblin_soldier", "고블린 병사", 50, 15, 5, "Normal", "iron_scrap,goblin_leather", 20],
    ["goblin_archer", "고블린 궁수", 40, 20, 2, "Fast", "wood,arrow", 25],
    ["goblin_shaman", "고블린 샤먼", 45, 25, 3, "Normal", "magic_material,potion", 35],
    ["goblin_warrior", "고블린 전사", 80, 30, 10, "Normal", "weapon,armor", 50],
    ["skeleton", "스켈레톤", 60, 20, 8, "Normal", "bone,ancient_relic", 40],
    ["skeleton_archer", "스켈레톤 궁수", 50, 25, 5, "Fast", "bone,arrow", 45],
    ["skeleton_warrior", "스켈레톤 전사", 100, 35, 15, "Normal", "steel_weapon,bone", 70],
    ["goblin_chief", "고블린 족장", 500, 50, 25, "Fast", "legendary_weapon,treasure", 500],
    ["ancient_golem", "고대 골렘", 1000, 80, 40, "Slow", "ancient_relic,special_material", 1000],
]
for col, header in enumerate(monster_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
for row_idx, row_data in enumerate(monster_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border

# Bosses sheet
ws3 = wb.create_sheet("Bosses")
boss_headers = ["BossID", "BossName", "PhaseCount", "PhaseHPThresholds", "Phase1Skills", "Phase2Skills", "Phase3Skills", "Rewards", "MonsterID"]
boss_data = [
    ["goblin_chief_boss", "고블린 족장", 3, "100,60,30", "triple_slash,jump_attack,summon", "spin_attack,poison_cloud", "rage,continuous_summon", "legendary_axe,treasure_chest,crown", "goblin_chief"],
]
for col, header in enumerate(boss_headers, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
for row_idx, row_data in enumerate(boss_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border

wb.save(os.path.join(output_dir, "05_Combat.xlsx"))
print(f"Created: {os.path.join(output_dir, '05_Combat.xlsx')}")

print("\n✅ Excel 파일 생성 완료!")
print(f"📁 저장 위치: {os.path.abspath(output_dir)}")
